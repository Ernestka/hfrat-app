from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.utils import timezone
from datetime import timedelta

from .models import ResourceReport, User, Facility, ResourceReportHistory
from .serializers import (
    ResourceReportSerializer,
    DashboardFacilityReportSerializer,
    AdminCreateUserSerializer,
    FacilitySerializer,
    AdminUserListSerializer,
)
from .permissions import ReporterOnly, MonitorOnly, MonitorOrAdmin, AdminOnly


@api_view(["GET"])
def health_check(request):
    # Include role hint for frontend routing
    role = None
    facility_info = None
    user = request.user if getattr(
        request, "user", None) and request.user.is_authenticated else None
    if user:
        if getattr(user, "role", None) == User.Role.ADMINISTRATOR:
            role = "ADMIN"
        elif user.is_staff or user.is_superuser:
            role = "ADMIN"
        else:
            role = getattr(user, "role", None)

        # Include facility info for reporters
        if role == "REPORTER" and hasattr(user, 'facility') and user.facility:
            facility_info = {
                "name": user.facility.name,
                "city": user.facility.city,
                "country": user.facility.country
            }

    response_data = {"status": "ok", "role": role}
    if facility_info:
        response_data["facility"] = facility_info
    return Response(response_data)


class ReporterResourceReportView(APIView):
    permission_classes = [IsAuthenticated, ReporterOnly]

    def _upsert(self, request):
        user = request.user
        if not getattr(user, "facility", None):
            return Response({"detail": "Assigned facility required for reporters."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = ResourceReportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        report, _ = ResourceReport.objects.update_or_create(
            facility=user.facility,
            defaults=serializer.validated_data,
        )

        # Save historical snapshot
        ResourceReportHistory.objects.create(
            facility=user.facility,
            icu_beds_available=report.icu_beds_available,
            ventilators_available=report.ventilators_available,
            staff_on_duty=report.staff_on_duty,
        )

        return Response(ResourceReportSerializer(report).data, status=status.HTTP_200_OK)

    def post(self, request):
        return self._upsert(request)

    def put(self, request):
        return self._upsert(request)


class MonitorDashboardView(APIView):
    permission_classes = [IsAuthenticated, MonitorOrAdmin]

    def get(self, request):
        queryset = ResourceReport.objects.select_related("facility").all()
        serializer = DashboardFacilityReportSerializer(queryset, many=True)
        return Response(serializer.data)


class MonitorExportDashboardView(APIView):
    """Export dashboard data to Excel"""
    permission_classes = [IsAuthenticated, MonitorOrAdmin]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime

        reports = ResourceReport.objects.select_related(
            "facility").all().order_by("facility__name")

        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Report"

        # Header styling
        header_fill = PatternFill(
            start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ["Facility Name", "City", "Country", "ICU Beds Available",
                   "Ventilators Available", "Staff on Duty", "Status", "Last Updated"]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows
        for report in reports:
            status = "CRITICAL" if report.icu_beds_available == 0 else "OK"
            row = [
                report.facility.name,
                report.facility.city,
                report.facility.country,
                report.icu_beds_available,
                report.ventilators_available,
                report.staff_on_duty,
                status,
                report.last_updated.strftime(
                    '%Y-%m-%d %H:%M:%S') if report.last_updated else "N/A",
            ]
            ws.append(row)

            # Color-code status cells
            last_row = ws.max_row
            status_cell = ws.cell(row=last_row, column=7)
            if status == "CRITICAL":
                status_cell.fill = PatternFill(
                    start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                status_cell.font = Font(color="b91c1c", bold=True)
            else:
                status_cell.fill = PatternFill(
                    start_color="ecfdf3", end_color="ecfdf3", fill_type="solid")
                status_cell.font = Font(color="166534", bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20

        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"dashboard_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class AdminCreateUserView(APIView):
    permission_classes = [IsAuthenticated, AdminOnly]

    def post(self, request):
        serializer = AdminCreateUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(serializer.to_representation(user), status=status.HTTP_201_CREATED)


class AdminFacilityView(APIView):
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request):
        facilities = Facility.objects.all().order_by("name")
        data = FacilitySerializer(facilities, many=True).data
        return Response(data)

    def post(self, request):
        serializer = FacilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AdminUserListView(APIView):
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request):
        users = User.objects.select_related("facility").order_by("username")
        data = AdminUserListSerializer(users, many=True).data
        return Response(data)


class AdminUserDetailView(APIView):
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request, user_id):
        try:
            user = User.objects.select_related("facility").get(id=user_id)
            data = AdminUserListSerializer(user).data
            return Response(data)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        from .serializers import AdminUpdateUserSerializer
        serializer = AdminUpdateUserSerializer(
            user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated_user = serializer.save()
        return Response(AdminUserListSerializer(updated_user).data)

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            # Prevent deletion of current user
            if user.id == request.user.id:
                return Response({"detail": "Cannot delete your own account."}, status=status.HTTP_400_BAD_REQUEST)
            user.delete()
            return Response({"detail": "User deleted successfully."}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


class AdminExportUsersView(APIView):
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request):
        from openpyxl import Workbook
        from django.http import HttpResponse
        from datetime import datetime

        users = User.objects.select_related("facility").order_by("username")

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Headers
        headers = ["ID", "Username", "Role", "Hospital", "City", "Country"]
        ws.append(headers)

        # Data rows
        for user in users:
            row = [
                user.id,
                user.username,
                user.role,
                user.facility.name if user.facility else "N/A",
                user.facility.city if user.facility else "N/A",
                user.facility.country if user.facility else "N/A",
            ]
            ws.append(row)

        # Create HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class MonitorTrendView(APIView):
    """API endpoint for 7-day historical trend data"""
    permission_classes = [IsAuthenticated, MonitorOrAdmin]

    def get(self, request):
        # Get facility_id from query params
        facility_id = request.query_params.get('facility_id')

        if not facility_id:
            return Response(
                {"detail": "facility_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            facility = Facility.objects.get(id=facility_id)
        except Facility.DoesNotExist:
            return Response(
                {"detail": "Facility not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get data from last 7 days
        seven_days_ago = timezone.now() - timedelta(days=7)
        history = ResourceReportHistory.objects.filter(
            facility=facility,
            timestamp__gte=seven_days_ago
        ).order_by('timestamp')

        # Group by day and get daily averages
        from django.db.models import Avg
        from django.db.models.functions import TruncDate

        daily_data = history.annotate(
            day=TruncDate('timestamp')
        ).values('day').annotate(
            avg_beds=Avg('icu_beds_available'),
            avg_vents=Avg('ventilators_available'),
            avg_staff=Avg('staff_on_duty'),
        ).order_by('day')

        # Format response
        trend_data = {
            "facility_id": facility.id,
            "facility_name": facility.name,
            "city": facility.city,
            "country": facility.country,
            "data": [
                {
                    "date": item['day'].strftime('%Y-%m-%d'),
                    "icu_beds": round(item['avg_beds'], 1),
                    "ventilators": round(item['avg_vents'], 1),
                    "staff": round(item['avg_staff'], 1),
                }
                for item in daily_data
            ]
        }

        return Response(trend_data)


class AdminPlatformStatsView(APIView):
    """Get platform-wide statistics for admin dashboard"""
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request):
        from django.db.models import Count, Sum, Avg

        # Count statistics
        total_facilities = Facility.objects.count()
        total_users = User.objects.count()
        total_reports = ResourceReport.objects.count()
        total_history_records = ResourceReportHistory.objects.count()

        # User role breakdown
        user_stats = User.objects.values('role').annotate(count=Count('role'))

        # Resource totals
        resource_totals = ResourceReport.objects.aggregate(
            total_beds=Sum('icu_beds_available'),
            total_vents=Sum('ventilators_available'),
            total_staff=Sum('staff_on_duty'),
            avg_beds=Avg('icu_beds_available'),
            avg_vents=Avg('ventilators_available'),
            avg_staff=Avg('staff_on_duty'),
        )

        # Critical facilities count
        critical_facilities = ResourceReport.objects.filter(
            icu_beds_available=0
        ).count()

        # Facilities by country
        facilities_by_country = Facility.objects.values('country').annotate(
            count=Count('id')
        ).order_by('-count')

        return Response({
            "overview": {
                "total_facilities": total_facilities,
                "total_users": total_users,
                "total_reports": total_reports,
                "total_history_records": total_history_records,
                "critical_facilities": critical_facilities,
            },
            "users_by_role": {item['role']: item['count'] for item in user_stats},
            "resources": {
                "total_beds": resource_totals['total_beds'] or 0,
                "total_ventilators": resource_totals['total_vents'] or 0,
                "total_staff": resource_totals['total_staff'] or 0,
                "avg_beds_per_facility": round(resource_totals['avg_beds'] or 0, 1),
                "avg_ventilators_per_facility": round(resource_totals['avg_vents'] or 0, 1),
                "avg_staff_per_facility": round(resource_totals['avg_staff'] or 0, 1),
            },
            "facilities_by_country": list(facilities_by_country),
        })


class AdminSettingsListView(APIView):
    """Get all system settings"""
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request):
        from .models import SystemSetting
        from .serializers import SystemSettingSerializer

        settings = SystemSetting.objects.all()
        serializer = SystemSettingSerializer(settings, many=True)
        return Response(serializer.data)

    def post(self, request):
        """Create a new setting"""
        from .models import SystemSetting
        from .serializers import SystemSettingSerializer

        serializer = SystemSettingSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AdminSettingsDetailView(APIView):
    """Get, update, or delete a specific setting"""
    permission_classes = [IsAuthenticated, AdminOnly]

    def get(self, request, setting_id):
        from .models import SystemSetting
        from .serializers import SystemSettingSerializer

        try:
            setting = SystemSetting.objects.get(pk=setting_id)
        except SystemSetting.DoesNotExist:
            return Response(
                {"error": "Setting not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = SystemSettingSerializer(setting)
        return Response(serializer.data)

    def put(self, request, setting_id):
        from .models import SystemSetting
        from .serializers import SystemSettingUpdateSerializer

        try:
            setting = SystemSetting.objects.get(pk=setting_id)
        except SystemSetting.DoesNotExist:
            return Response(
                {"error": "Setting not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = SystemSettingUpdateSerializer(
            setting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            # Return full setting data
            from .serializers import SystemSettingSerializer
            full_serializer = SystemSettingSerializer(setting)
            return Response(full_serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, setting_id):
        from .models import SystemSetting

        try:
            setting = SystemSetting.objects.get(pk=setting_id)
        except SystemSetting.DoesNotExist:
            return Response(
                {"error": "Setting not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        setting.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminSettingsInitializeView(APIView):
    """Initialize default system settings"""
    permission_classes = [IsAuthenticated, AdminOnly]

    def post(self, request):
        from .models import SystemSetting

        default_settings = [
            {
                "key": "critical_icu_beds_threshold",
                "value": "5",
                "description": "Minimum ICU beds before facility is marked CRITICAL",
                "setting_type": "THRESHOLD"
            },
            {
                "key": "critical_ventilators_threshold",
                "value": "3",
                "description": "Minimum ventilators before facility is marked CRITICAL",
                "setting_type": "THRESHOLD"
            },
            {
                "key": "critical_staff_threshold",
                "value": "10",
                "description": "Minimum staff on duty before facility is marked CRITICAL",
                "setting_type": "THRESHOLD"
            },
            {
                "key": "alert_notification_enabled",
                "value": "true",
                "description": "Enable or disable alert notifications",
                "setting_type": "ALERT"
            },
            {
                "key": "dashboard_refresh_interval",
                "value": "60",
                "description": "Dashboard auto-refresh interval in seconds",
                "setting_type": "GENERAL"
            },
        ]

        created_count = 0
        existing_count = 0

        for setting_data in default_settings:
            _, created = SystemSetting.objects.get_or_create(
                key=setting_data["key"],
                defaults={
                    "value": setting_data["value"],
                    "description": setting_data["description"],
                    "setting_type": setting_data["setting_type"],
                    "updated_by": request.user,
                }
            )
            if created:
                created_count += 1
            else:
                existing_count += 1

        return Response({
            "message": "Settings initialization complete",
            "created": created_count,
            "existing": existing_count,
            "total": len(default_settings)
        })


class PublicSettingsView(APIView):
    """Get public system settings (thresholds) for dashboard display"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import SystemSetting

        # Get threshold settings only
        thresholds = SystemSetting.objects.filter(setting_type='THRESHOLD')

        result = {}
        for setting in thresholds:
            try:
                # Try to convert to int for numeric thresholds
                result[setting.key] = int(setting.value)
            except ValueError:
                result[setting.key] = setting.value

        return Response(result)
